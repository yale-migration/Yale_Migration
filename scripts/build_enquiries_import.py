#!/usr/bin/env python3
"""
Build the ENQUIRIES import from the client's `DATA SHEET.xlsx` cold-call log.  (D-327)

    python3 scripts/build_enquiries_import.py            # report only, writes nothing
    python3 scripts/build_enquiries_import.py --write    # writes ../client-data/enquiries-import.csv

Reads from and writes to `../client-data/`, OUTSIDE the repo. Prints counts and
anonymised samples only — never a client name, never a phone number.

================================ THE DATE REPAIR ================================
🔴 47% of the dates in this file are WRONG IN THE FILE, day and month transposed.
Import them as they stand and roughly half of Yale's enquiry log is off by up to
five months — and SOP-CI-001's follow-up cadence (7 days, then 30) fires against
those dates. A June enquiry filed as January is never chased; a fresh one filed in
the future is never chased either.

How we know, rather than suspect. Of the 566 populated date cells:

    301 are TEXT      '26/06/2026'   — day > 12 in 301 of them          = 100%
    265 are DATETIME  2026-01-07     — day > 12 in   0 of them          =   0%

A perfect split on "is the day greater than 12" cannot happen by chance. It is the
signature of Excel parsing with a US locale:

    '7/1/2026'  -> both parts <= 12 -> read as m/d -> JULY 1     -> stored as a date
    '26/6/2026' -> 26 is no month   -> parse fails -> left alone -> stayed text

So every datetime-typed cell is an Australian d/m string that Excel read as m/d.

Confirmed by an independent signal that was never used to build the hypothesis —
impossible future dates in a call log that ends in August:

    as stored          2026-01-07 .. 2026-11-08     55 dates in the future
    day/month swapped  2026-07-01 .. 2026-08-11      0 dates in the future

55 -> 0, and the repaired range lines up exactly with the text cells' own range.

⛔ The repair is applied ONLY to datetime-typed cells. Text cells are already correct
and touching them would break what works.

============================ WHAT IS DELIBERATELY BLANK =========================
`Status`  SOP-CI-001 step 10B gives the vocabulary — Not Proceeding / Pending
          Decision / Lost Lead — and the ENQUIRIES dropdown already carries it. But
          it is a vocabulary from a PROCESS DIAGRAM, and **no column in DATA SHEET
          records it**. Searched all four client workbooks: those three phrases
          appear in none of them. The remarks hold 'call back' (22), 'no response'
          (3), 'follow up' (3), 'not interested' (1) — and mapping 'call back' onto
          'Pending Decision' would be us inventing a migration agency's lead status
          and then feeding it to an automation. Left blank. The raw remark is
          carried into Notes so a human can set it in seconds.

`Channel` ⛔ ANSWERED 18 Aug — AND THE ANSWER KILLED MY DEFAULT. I had set this to
          'Phone' because the column is titled 'Phone Number' and the remarks are full
          of 'call back'. Rey (Reyward Jake Gamol, consultant) replied:

              "Inquiries usually come from both whatsapp and social media accounts"

          Not phone. Two channels, and no column in the file distinguishes them
          row-by-row — so stamping either one on 621 rows is wrong for roughly half.
          Default is now BLANK, which is a STRONGER position than the blank I started
          with: then it was uncertainty, now it is positive evidence of a mix.

          🔑 The default was live for about four hours. Had nobody asked, "Phone" would
          have sat in 621 cells and become the truth by attrition. The label that said
          ASSUMPTION every time the script ran is the only reason it was still cheap to
          undo. Defaults must announce themselves.

          --channel WhatsApp still works if they later say WhatsApp dominates.
"""
import argparse
import collections
import csv
import datetime
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "..", "..", "client-data", "DATA SHEET.xlsx")
OUT = os.path.join(HERE, "..", "..", "client-data", "enquiries-import.csv")

# Must match ENQUIRY_HEADERS in setup_master_sheet.gs exactly, in order.
HEADERS = ["Date", "Name", "Phone", "Email", "Channel", "Visa Interest",
           "Location", "Assigned To", "Status", "Follow-up Due", "Notes"]

# `Assigned To` is a dropdown built with setAllowInvalid(false) — a value outside
# this list is REJECTED by the cell, so every name has to land on one of them.
ROSTER = ["Robinder", "Inder", "Gayatri", "Priyanka", "Fiza", "RJ", "Star",
          "Rey", "Cristelle", "Unassigned"]
STAFF_MAP = {
    "inder": "Inder", "indert": "Inder",      # 'indert' is a typo, 1 row
    "rj": "RJ", "gayatri": "Gayatri", "priyanka": "Priyanka", "fiza": "Fiza",
    "robin": "Robinder",                       # their shorthand for Robinder
    "rey": "Rey", "star": "Star", "cristelle": "Cristelle",
    "none": "Unassigned", "": "Unassigned",
}

TODAY = datetime.date.today()


def load_rows():
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is not installed:  pip3 install openpyxl")
    if not os.path.exists(SRC):
        sys.exit("cannot find %s — client data lives OUTSIDE the repo" % SRC)
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Sheet1"]
    out = []
    for r in range(2, ws.max_row + 1):          # row 1 is the header (and is damaged)
        vals = [ws.cell(r, c).value for c in range(1, 7)]
        if any(v is not None and str(v).strip() for v in vals):
            out.append((r, vals))
    wb.close()
    return out


def repair_date(v):
    """-> (date|None, tag). See the module docstring for why the swap is correct."""
    if v is None or not str(v).strip():
        return None, "missing"
    if isinstance(v, datetime.datetime):
        try:
            return datetime.date(v.year, v.day, v.month), "repaired"   # swap d/m
        except ValueError:
            # day>12 in a datetime cell would contradict the whole finding. Do not
            # silently keep it — say so, because it means the file has changed shape.
            return v.date(), "datetime-unswappable"
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        try:
            return datetime.date(y, mo, d), "text-ok"                  # already d/m
        except ValueError:
            return None, "text-invalid"
    return None, "unparseable"


def clean_phone(p):
    """Digits and a leading +. Keeps international numbers; drops Excel's decimals."""
    s = str(p or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[^\d+]", "", s)
    return s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="write the CSV (default: report only)")
    ap.add_argument("--channel", default="",
                    choices=["Phone", "WhatsApp", "Walk-in", "Email", "Website",
                             "Referral", "SMS", "Facebook", "Instagram", ""],
                    help="Channel to stamp on every row. Default BLANK: the team says "
                         "enquiries are a WhatsApp/social-media mix and nothing in the "
                         "file separates them row-by-row (D-330).")
    args = ap.parse_args()

    rows = load_rows()
    tags = collections.Counter()
    staff_unmapped = collections.Counter()
    out, seen_phone, dupes, no_phone_no_name = [], {}, 0, 0
    future = []

    for src_row, (d, name, phone, staff, enquiry, remarks) in rows:
        date, tag = repair_date(d)
        tags[tag] += 1
        if date and date > TODAY:
            future.append((src_row, date))

        name = str(name or "").strip()
        phone = clean_phone(phone)
        enquiry = str(enquiry or "").strip()
        remarks = str(remarks or "").strip()

        if not name and not phone:
            no_phone_no_name += 1
            continue

        raw_staff = str(staff or "").strip().lower()
        if raw_staff in STAFF_MAP:
            assigned = STAFF_MAP[raw_staff]
            extra = ""
        else:
            assigned = "Unassigned"
            staff_unmapped[raw_staff] += 1
            # Never discard what they typed. It goes to Notes for a human to read.
            extra = ' | staff column said: "%s"' % str(staff).strip()

        # Dedupe on phone. Keep the EARLIEST enquiry: SOP-CI-001's cadence runs from
        # first contact, so the first row is the one the follow-up clock belongs to.
        key = phone or ("name:" + name.lower())
        if key in seen_phone:
            dupes += 1
            prev = out[seen_phone[key]]
            if date and (not prev[0] or str(date) < prev[0]):
                prev[0] = str(date)
            continue

        note = "Imported from DATA SHEET row %d on %s" % (src_row, TODAY)
        if tag == "repaired":
            note += " | date day/month repaired (D-327)"
        if remarks:
            note += " | " + remarks
        note += extra

        seen_phone[key] = len(out)
        out.append([
            str(date) if date else "",
            name,
            phone,
            "",                 # Email — not in this file
            args.channel,       # ASSUMED — see the docstring, reversible in one step
            enquiry,            # Visa Interest — their own words, unaltered
            "",                 # Location — not in this file
            assigned,
            "",                 # Status — deliberately blank, see the docstring
            "",                 # Follow-up Due — M6 computes this
            note,
        ])

    print("=== SOURCE ===")
    print("  non-empty rows in DATA SHEET ....... %d" % len(rows))
    print("  dropped, no name AND no phone ...... %d" % no_phone_no_name)
    print("  duplicates merged (same phone) ..... %d" % dupes)
    print("  ROWS TO IMPORT ..................... %d" % len(out))

    print("\n=== DATES ===")
    for k, v in tags.most_common():
        flag = "  <-- day/month swapped" if k == "repaired" else ""
        print("  %-22s %4d%s" % (k, v, flag))
    print("  dates still in the future .......... %d %s"
          % (len(future), [str(d) for _, d in future[:3]]))
    if future:
        print("  ⚠️  a call log cannot have future dates. These are typos in the")
        print("      SOURCE file, not parse errors — check them by hand before import.")

    print("\n=== ASSIGNED TO ===")
    got = collections.Counter(r[7] for r in out)
    for k in ROSTER:
        if got.get(k):
            print("  %-12s %4d" % (k, got[k]))
    if staff_unmapped:
        print("  values that matched nobody on the roster -> Unassigned,")
        print("  and preserved verbatim in Notes:")
        for k, v in staff_unmapped.most_common(8):
            print('      "%s" x%d' % (k[:44] if k else "(blank)", v))

    print("\n=== JUDGEMENTS AND BLANKS ===")
    print("  Channel  — %r. The team confirmed enquiries are a WhatsApp/social-media"
          % (args.channel or "(blank)"))
    print("             mix (Rey, 18 Aug) and no column separates them per row, so blank")
    print("             is the honest value. --channel WhatsApp if that changes.")
    print("  Status   — BLANK. SOP-CI-001's vocabulary is real but NO column in any of")
    print("             the four client workbooks records it. Not invented here.")
    print("  Email    — BLANK. Not present in this file at all.")

    if args.write:
        with open(OUT, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(HEADERS)
            w.writerows(out)
        print("\nWROTE %s  (%d rows)" % (os.path.normpath(OUT), len(out)))
        print("⛔ That file holds client PII. It lives outside the repo. Do not commit it.")
    else:
        print("\n(report only — pass --write to produce the CSV)")


if __name__ == "__main__":
    main()
