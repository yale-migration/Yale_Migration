#!/usr/bin/env python3
"""Repo hygiene gate — run before every commit and at session end.

    python3 scripts/repo_hygiene.py

Exists because on 15 Aug I wrote a real client's name into a code comment and
into DECISIONS.md while documenting a finding about that very client, and only
caught it with an ad-hoc grep afterwards.  CLAUDE.md says "NO SECRETS in this
repo — no keys, no passwords, no client PII."  A rule with no check is a
preference.

It also caught a second thing by accident: the client's live lodgement list
contains a row literally named SAMPLE, which had already reached row 1 of the
pilot import CSV and would have become a real OneDrive client folder.

Client names are NOT hardcoded here — that would put them in the repo.  They are
read at runtime from ../client-data/, which lives outside the repo.

Design notes, both learned the hard way:
  * Match on the FULL name, never a single token.  Our demo data uses "Sharma"
    and so does a real client; flagging that trains everyone to ignore the check.
  * A tracked spreadsheet is only a problem if it actually holds personal data.
    The two fee workbooks in docs/ are price lists and a blank invoice template.
"""

import os
import re
import subprocess
import sys
import warnings

warnings.filterwarnings("ignore")

os.chdir(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

TEXT_EXT = ("md", "py", "gs", "sh", "json", "js", "ts", "txt")
JUNK = {"SAMPLE", "TEST", "DEMO", "EXAMPLE", "NAME", "N/A", "TBA", "XXX"}
DATA = os.path.join("..", "client-data")

EMAIL = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE = re.compile(r"\b(?:\+?61|0)[24578][0-9]{8}\b")

SECRETS = [
    (r"\bBSB[ :.-]*[0-9]{3}[ -]?[0-9]{3}\b",              "BSB with digits"),
    (r"\bSWIFT[ :.-]*[A-Z]{4}AU[0-9A-Z]{2}\b",            "SWIFT code"),
    (r"\baccount ?(?:number|no)[ :.-]*[0-9]{6,}",         "account number"),
    (r"\b(?:api[_ ]?key|secret[_ ]?key|access[_ ]?token)\b *[:=] *['\"]?[A-Za-z0-9_\-]{12,}",
                                                          "API key / token"),
    (r"\bBearer +[A-Za-z0-9._\-]{20,}",                   "bearer token"),
    (r"-----BEGIN [A-Z ]*PRIVATE KEY",                    "private key"),
    (r"\b(?:password|passwd|pwd)\b *[:=] *['\"]?[^\s'\"<*{$\[]{4,}",
                                                          "literal password"),
]
# Lines that legitimately DISCUSS credentials rather than containing one.
ABOUT = re.compile(
    r"column|field|header|rule|exclud|never|plaintext|rotat|manager|redact"
    r"|placeholder|example|<[^>]*>|\.\.\.|xxx|D-306|no-secrets|hygiene",
    re.I)

fail = False


def tracked_text_files():
    out = subprocess.run(["git", "ls-files"], capture_output=True, text=True).stdout
    return [f for f in out.split("\n")
            if f and f.rsplit(".", 1)[-1].lower() in TEXT_EXT and os.path.exists(f)]


def head(n, title):
    print("\n=== %d. %s ===" % (n, title))


# ---------------------------------------------------------------- 1. secrets
head(1, "credential-shaped strings")
hits = 0
for f in tracked_text_files():
    if f == os.path.join("scripts", "repo_hygiene.py"):
        continue                       # the patterns themselves live here
    try:
        lines = open(f, encoding="utf-8", errors="ignore").read().split("\n")
    except OSError:
        continue
    for i, line in enumerate(lines, 1):
        for rx, label in SECRETS:
            if re.search(rx, line, re.I) and not ABOUT.search(line):
                print("  %s:%d  %s" % (f, i, label))
                hits += 1
print("  clean" if not hits else "  ^^ FAIL")
fail |= bool(hits)

# ------------------------------------------------------- 2. client full names
head(2, "real client names leaking in from the workbooks")
names = []
src = os.path.join(DATA, "YALE BRISBANE OFFICE WORK.xlsx")
if not os.path.isdir(DATA):
    print("  SKIP — %s not present" % DATA)
elif not os.path.exists(src):
    print("  SKIP — source workbook not present")
else:
    import openpyxl
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    for tab in ("LODGEMENT JULY TO PRESENT",):
        if tab not in wb.sheetnames:
            continue
        rows = list(wb[tab].iter_rows(values_only=True))
        hdr = [str(c).strip() if c else "" for c in rows[0]]
        if "NAME" not in hdr:
            continue
        j = hdr.index("NAME")
        for r in rows[1:]:
            if j < len(r) and isinstance(r[j], str) \
                    and r[j].strip().upper() not in JUNK:
                toks = [t.upper() for t in re.sub(r"[^A-Za-z ]", " ", r[j]).split()
                        if len(t) >= 4]
                if len(toks) >= 2:
                    names.append((r[j].strip(), toks))
    wb.close()

    blobs = {}
    for f in tracked_text_files():
        try:
            blobs[f] = open(f, encoding="utf-8", errors="ignore").read().upper()
        except OSError:
            pass
    leaked = 0
    for full, toks in names:
        for f, txt in blobs.items():
            if all(re.search(r"\b" + re.escape(t) + r"\b", txt) for t in toks):
                print("  LEAKED: %s  ->  %s" % (full, f))
                leaked += 1
    print("  clean — none of the %d active-client full names appears in any "
          "tracked file" % len(names) if not leaked else "  ^^ FAIL")
    fail |= bool(leaked)

# ------------------------------------------------ 3. tracked spreadsheet data
head(3, "tracked spreadsheets must not hold personal data")
sheets = [f for f in subprocess.run(["git", "ls-files"], capture_output=True,
                                    text=True).stdout.split("\n")
          if f.lower().endswith((".xlsx", ".xls", ".csv")) and os.path.exists(f)]
if not sheets:
    print("  none tracked")
else:
    import openpyxl
    bad = 0
    for f in sheets:
        n = 0
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            for tab in wb.sheetnames:
                for i, row in enumerate(wb[tab].iter_rows(values_only=True)):
                    if i > 300:
                        break
                    for c in row:
                        if isinstance(c, str) and (EMAIL.search(c) or PHONE.search(c)):
                            n += 1
            wb.close()
        except Exception as e:
            print("  ?  unreadable: %s (%s)" % (f, str(e)[:40]))
            continue
        if n >= 3:
            print("  PERSONAL DATA: %s — %d email/phone cells" % (f, n))
            bad += 1
        else:
            print("  ok, reference only (%d hits): %s" % (n, f))
    fail |= bool(bad)

# ------------------------------------------- 4. Apps Script global-name collisions
#
# 🔴 Apps Script shares ONE global scope across every .gs file in a project. Two files
# declaring the same top-level `var` or `function` collide, and whichever loads last
# wins — non-deterministically, from our point of view.
#
# This exists because of a real one (D-326): `CF_HEADER` was 'Checklist Filed' in
# setup_m4_checklist_map.gs and 'Chase Flag' in add_chase_flag_column_ae.gs. The run
# log printed `Column "Chase Flag" already present.` from a function that looks for
# 'Checklist Filed'. It was harmless only by luck — the column it checked happened to
# exist, so it early-returned instead of writing 'Chase Flag' over MASTER's column Y.
#
# DIFFERENT values = FAIL. IDENTICAL values = warn (still fragile: one edit apart).
head(4, "Apps Script global-name collisions across scripts/*.gs")
import collections
_decl = collections.defaultdict(dict)     # name -> {file: value-or-None}
_gsdir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
_gs = sorted(f for f in os.listdir(_gsdir) if f.endswith(".gs"))
for _f in _gs:
    for _line in open(os.path.join(_gsdir, _f), encoding="utf-8"):
        _m = re.match(r"var\s+([A-Za-z_$][\w$]*)\s*=\s*(.+?);\s*(?://.*)?$", _line)
        if _m:
            _decl[_m.group(1)].setdefault(_f, _m.group(2).strip())
            continue
        _m = re.match(r"function\s+([A-Za-z_$][\w$]*)", _line)
        if _m:
            _decl[_m.group(1)].setdefault(_f, None)

_hard, _soft = [], []
for _name, _where in sorted(_decl.items()):
    if len(_where) < 2:
        continue
    _vals = set(_where.values())
    (_hard if len(_vals) > 1 else _soft).append((_name, _where))

for _name, _where in _soft:
    print("  warn  %s — same value in %s" % (_name, ", ".join(sorted(_where))))
for _name, _where in _hard:
    print("  FAIL  %s — DIFFERENT values, last file loaded wins:" % _name)
    for _f, _v in sorted(_where.items()):
        print("          %-34s %s" % (_f, _v if _v is not None else "(function)"))
if not _hard and not _soft:
    print("  clean — no shared top-level names")
fail |= bool(_hard)

# ------------------------------------------------------------- 5. remote/push
head(5, "remote and unpushed state")
remote = subprocess.run(["git", "remote", "get-url", "origin"],
                        capture_output=True, text=True).stdout.strip() or "none"
ahead = subprocess.run(["git", "rev-list", "--count", "@{u}..HEAD"],
                       capture_output=True, text=True).stdout.strip() or "?"
print("  remote:   " + remote)
print("  unpushed: %s commits" % ahead)
COMPANY = ("BrandRadar-AI", "Roar-AI-Labs", "Apex-AI-Clients")
if remote != "none" and not any(o.lower() in remote.lower() for o in COMPANY):
    print("  ⚠️  NOT a company GitHub org (%s)." % ", ".join(COMPANY))
    print("      This repo documents client data. Confirm with the user")
    print("      before pushing — org policy requires it.")

print()
print("HYGIENE PASS" if not fail else "HYGIENE FAIL — fix before committing")
sys.exit(1 if fail else 0)
