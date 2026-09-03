#!/usr/bin/env python3
"""
Build the MASTER import from the team's RETURNED client list.          (D-330)

    python3 scripts/build_master_import.py            # report only, writes nothing
    python3 scripts/build_master_import.py --write    # -> ../client-data/master-import.csv

Two sources, joined on client name:
  1. `2026-08-18_CLIENT-LIST-TO-UPDATE_returned.xlsx` — what the team filled in on
     18 Aug: team, consultant, email, office, surnames, skills authority.
  2. `YALE BRISBANE OFFICE WORK.xlsx` -> `LODGEMENT JULY TO PRESENT` — their live tab,
     which still owns STATUS, CURRENT VISA and VISA EXPIRATION.

The join was verified before this script was written: 39 of 40 returned names match a
row in the original exactly. Two originals are absent from the returned list — the
`SAMPLE` test row and one duplicate — which is what we expected them to drop.

⛔ Report-only by default. The CSV holds 40 real people. It is written to
`../client-data/`, outside the repo, and must never be committed.

============================ WHAT THIS SCRIPT REFUSES TO DO =========================
* ~~It does not repair the `gmil.com` email.~~ 🔓 **REFUSAL LIFTED 22 Aug.** The rule was
  right when written — "almost certainly a typo" is not a basis for inventing a real
  client's address. But the human has now decided: RJ, in writing, *"gmail.com always is
  the correct one."* The premise of the refusal was that WE would be guessing. We are not
  guessing any more, so it is repaired, noted on the row, and counted. See CONFIRMED_FIX.
  🔑 A guard exists to stop us acting without authority. When the authority arrives, the
  guard comes down — it does not become a tradition.
* It does not guess the 4 missing 485 skills authorities. The team left those blank.
  Without one, M4 route B stamps NEEDS REVIEW — which is the correct outcome, not a
  bug to paper over.
* It does not assign a consultant where the sheet says `NO LONGER CLIENT`. Those rows
  are HELD BACK entirely (see below).
"""
import argparse, collections, csv, datetime, glob, os, re, sys, warnings
warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
CD   = os.path.join(HERE, "..", "..", "client-data")
# 🔑 THE NEWEST RETURN WINS, and it is resolved at runtime rather than pinned.
# The team returned a second pass on 25 Aug (D-403): the rows 22/23 duplicate is
# gone, 10 contact numbers and 11 "anyone else on the application" answers
# arrived, and the four `<-- needed` skills authorities are filled. Reading the
# 18 Aug file would have quietly rebuilt the import from superseded data — and
# an import built from a stale source looks exactly like one built from a fresh
# one. ⛔ Sorted by DATE PREFIX, so a new return is picked up by dropping it in.
_RETURNS = sorted(glob.glob(os.path.join(CD, "20??-??-??_CLIENT-LIST-TO-UPDATE_returned*.xlsx")))
if not _RETURNS:
    raise SystemExit("❌ no CLIENT-LIST-TO-UPDATE return found in %s — refusing to guess" % CD)
RET  = _RETURNS[-1]
ORIG = os.path.join(CD, "YALE BRISBANE OFFICE WORK.xlsx")

# 🔴 ORIG IS A POINT-IN-TIME EXPORT OF A SHEET THEY EDIT DAILY. (D-430)
#
# `LODGEMENT: JULY TO PRESENT` lives in that workbook and the client works in it.
# On 3 Sep RJ said "I also added the contact number and email of the client in
# the July to present lodgment sheet" — while our copy was **19 days old**.
# Rebuilding then would have used the stale version and reported as missing the
# exact fields he had just filled in.
#
# ⛔ Nothing about a stale export looks wrong. It opens, it parses, every row is
# there. This is the same shape as D-404, where the importer read the older of
# two returns. That one was fixed by resolving the newest file at runtime; this
# one cannot be, because there is only ever one export and its age is invisible.
#
# So: refuse to run quietly on an old copy. The permanent fix is to stop reading
# an export at all — the dashboard sync reads the live tab over the Sheets API.
_ORIG_MAX_AGE_DAYS = 7

def _warn_if_stale(path, label, max_days):
    if not os.path.exists(path):
        return
    age = (datetime.datetime.now()
           - datetime.datetime.fromtimestamp(os.path.getmtime(path))).days
    if age > max_days:
        print("")
        print("  " + "=" * 68)
        print("  🔴 STALE SOURCE — %s is %d days old (limit %d)." % (label, age, max_days))
        print("     %s" % os.path.basename(path))
        print("     The client EDITS this sheet. Re-export it before trusting this run:")
        print("     open the workbook -> File -> Download -> Microsoft Excel (.xlsx)")
        print("     -> replace the file in client-data/, keeping the same name.")
        print("  " + "=" * 68)
        print("")
OUT  = os.path.join(CD, "master-import.csv")

# 🔴 THE LIVE GOOGLE SHEET AND THE .xlsx EXPORT DISAGREE ON THIS TAB NAME (D-337).
# Live:   'LODGEMENT: JULY TO PRESENT'   (colon)
# Export: 'LODGEMENT JULY TO PRESENT'    (Excel forbids ':' and renamed it silently)
# Anything pointed at the LIVE sheet must use the colon form or it finds no tab —
# and finds it without erroring, which is worse. Accept either, everywhere.
LODGEMENT_TAB_NAMES = ['LODGEMENT: JULY TO PRESENT', 'LODGEMENT JULY TO PRESENT']
GOOGLE_SHEET_IDS = {
    'YALE BRISBANE OFFICE WORK': '1NbaxgzHIiUM1yas1B3lt21ycNKyufPxXTxZPP0wamLI',
    'REYWARD JAKE M GAMOL-2026': '1_YDeb7iwHQr0c3MGKp0jp8MMyqBzqlr7sz36u8Qn4pc',
    'STUDENTS':                  '1XlnqEi42ZJNu3_vwNN8WgKcCk4zlzWyCyRQ9We_V9_A',
}

MASTER_HEADERS = ['Client Code','Their Client ID','Full Name','Party 2 Name','Contact Number',
 'Email Address','Location','Visa Type','Visa Variant','Office','Team','Assigned Consultant',
 'Processing Stage','Visa Outcome','Grant Date','Visa Expiry','Refusal Reason','Last Contact',
 'Next Follow-up Due','Date Added','Source','Folder URL','Notes','Skills Authority','Checklist Filed']

TEAM   = {'F': 'FILIPINO', 'I': 'INDIAN'}
# The sheet is upper-case and uses their shorthand. The dropdown is setAllowInvalid(false),
# so anything not landing on a roster value is REJECTED by the cell.
STAFF  = {'ROBIN':'Robinder','INDER':'Inder','RJ':'RJ','STAR':'Star','REY':'Rey',
          'PRIYANKA':'Priyanka','FIZA':'Fiza','GAYATRI':'Gayatri','CRISTELLE':'Cristelle'}
# Q15, answered 18 Aug: "Pending means not yet drafted and lodged."
STAGE  = {'LODGED':('Lodged','Pending'), 'DRAFTED':('Ready for Lodgement','Pending'),
          'PENDING':('Documents Pending','Pending'), 'WITHDRAWN':('Closed','Withdrawn')}
SKILLS = {'ACECQA','TRA','VETASSESS','Not required (Bachelor/Masters)','Engineers Australia'}
MAPPED = {'485','500','482','SBS','Nomination','407','820/801','189','190','491','494','802','101','417'}
# 🔓 Corrections the CLIENT has confirmed in writing. Each one carries its source, because
# the difference between a repair and an invention is entirely who authorised it.
CONFIRMED_FIX = {
    'domain': {
        # RJ, 22 Aug 2026: "gmail.com always is the correct one." One row, `gmil.com`.
        'gmil.com': ('gmail.com', 'RJ 22 Aug: "gmail.com always is the correct one"'),
    },
}

DEAD   = re.compile(r'no longer (a )?client', re.I)

# ══════════════════════════════════════════════════════════════════════════════
# LOCKED-COLUMN VALIDATION  (D-353, added 22 Aug 2026)
# ══════════════════════════════════════════════════════════════════════════════
# 🔴 Three times now, a value has been written into a column whose dropdown is
# `setAllowInvalid(false)` and would have REJECTED it, silently, at paste time:
#
#   D-138  SBS + Nomination      — "every sponsorship matter was a dead end"
#   A-33   GOPI                  — a consultant nobody could assign a client to
#   D-353  Citizenship           — 2 rows of this very import, plus 1 PARTNER VISA
#
# Someone thought about it for consultants (STAFF, above) and for nothing else.
# This closes the CLASS: every value bound for a locked column is checked against
# that column's own list before a CSV is written.
#
# 🔑 The lists are PARSED OUT OF setup_master_sheet.gs, never copied here. That
# file builds the sheet, so it is the only thing that knows what the sheet allows.
# A second copy would drift, and a validator that drifts is worse than none — it
# reports PASS against a schema nobody is using.
GS_SETUP = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'setup_master_sheet.gs')

# 🔴 MASTER'S DROPDOWNS ARE NOT ALL IN ONE FILE, AND THE GATE MISSED ONE. (D-403)
#
# `setup_master_sheet.gs` builds columns A–W and declares nine locked lists in
# MASTER_DROPDOWNS. Column X `Skills Authority` was added LATER by its own
# script, with its own `SA_VALUES` and its own `setAllowInvalid(false)` — so for
# every day this gate has existed it has validated nine of ten locked columns
# and reported PASS. **A gate with a silent blind spot reads exactly like a gate
# that passed.**
#
# Found when the client returned their list with the four `<-- needed` skills
# authorities filled in as FREE TEXT — "acecqa" lowercase, "Bachelor Degree -no
# skills assessment", "n/a". Every one of those is refused by the cell in
# silence. That is the FOURTH appearance of this bug class (LESSONS § 3).
#
# ⛔ Any future locked column added by its own script must be registered here.
EXTRA_DROPDOWN_SOURCES = [
    # (file, variable name, the MASTER header it locks)
    ('add_skills_authority_column.gs', 'SA_VALUES', 'Skills Authority'),
]

def locked_columns():
    """{header name: set(allowed values)} straight from the sheet builders."""
    src = open(GS_SETUP, encoding='utf-8').read()
    block = re.search(r'var MASTER_DROPDOWNS\s*=\s*\{(.*?)\n\};', src, re.S)
    if not block:
        raise SystemExit('❌ cannot find MASTER_DROPDOWNS in %s — refusing to guess' % GS_SETUP)
    body = re.sub(r'//[^\n]*', '', block.group(1))          # strip trailing comments
    out = {}
    for col, arr in re.findall(r'(\d+)\s*:\s*\[(.*?)\]', body, re.S):
        idx = int(col) - 1                                   # 1-based column -> 0-based header
        if idx < len(MASTER_HEADERS):
            out[MASTER_HEADERS[idx]] = set(re.findall(r"'([^']*)'", arr))

    here = os.path.dirname(os.path.abspath(__file__))
    for fname, var, header in EXTRA_DROPDOWN_SOURCES:
        path = os.path.join(here, fname)
        if not os.path.exists(path):
            raise SystemExit('❌ %s is missing — it defines the %s dropdown. Refusing to '
                             'validate against a schema I cannot read.' % (fname, header))
        m = re.search(r'var\s+%s\s*=\s*\[(.*?)\];' % re.escape(var), open(path, encoding='utf-8').read(), re.S)
        if not m:
            raise SystemExit('❌ cannot find %s in %s — refusing to guess' % (var, fname))
        vals = set(re.findall(r"'([^']*)'", re.sub(r'//[^\n]*', '', m.group(1))))
        if not vals:
            raise SystemExit('❌ %s in %s parsed to an EMPTY set — that would validate '
                             'everything and prove nothing' % (var, fname))
        out[header] = vals
    return out

# Their sheet's wording -> the dropdown's wording. ⛔ Only ever a RENAME of the
# same thing. Never a guess about what a client's visa actually is.
VISA_ALIAS = {
    'PARTNER VISA': '820/801',      # asked them to change the cell; they did not, and
                                    # it is our job to normalise, not their job to retype
}

# 🔑 D-352, promised to RJ in writing on 22 Aug: every imported client is stamped
# as already having had their checklist, because Yale sent them all by hand. M4's
# trigger is `Y notexist`, so a value here switches the checklist step off for
# this group and only this group. Client 39 onwards runs normally.
PRESTAMP_CHECKLIST = 'SENT BY YALE BEFORE IMPORT'

TODAY  = datetime.date.today()
norm   = lambda s: ' '.join(str(s or '').upper().split())


def sheets():
    import openpyxl
    for p in (RET, ORIG):
        if not os.path.exists(p):
            sys.exit("missing %s — client data lives OUTSIDE the repo" % os.path.basename(p))
    w = openpyxl.load_workbook(RET, data_only=True)["CLIENT LIST TO UPDATE"]
    ret = [[('' if w.cell(r, c).value is None else str(w.cell(r, c).value).strip())
            for c in range(1, 12)] for r in range(2, w.max_row + 1)]
    ret = [x for x in ret if x[0]]
    wb = openpyxl.load_workbook(ORIG, read_only=True, data_only=True)
    tab = next((t for t in LODGEMENT_TAB_NAMES if t in wb.sheetnames), None)
    if not tab:
        sys.exit("no lodgement tab found. Looked for: %s. Sheet has: %s"
                 % (LODGEMENT_TAB_NAMES, wb.sheetnames))
    rows = list(wb[tab].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    orig = {}
    for r in rows[1:]:
        nm = r[idx['NAME']] if idx.get('NAME') is not None and idx['NAME'] < len(r) else None
        if nm and norm(nm) not in orig:
            orig[norm(nm)] = {h: (r[i] if i < len(r) else None) for h, i in idx.items()}
    return ret, orig


def split_visa(raw):
    """'485 Dependent' -> ('485','Dependent').  '600  (no checklist yet)' -> ('600','')."""
    # 🔴 A ROUND-TRIP THROUGH GOOGLE SHEETS TURNS 500 INTO 500.0. (D-404)
    #
    # openpyxl hands back a float for any numeric cell, and the 25 Aug return
    # came back through Sheets, so 32 of 40 visa types arrived as floats. The
    # locked dropdown holds '500', not '500.0', and `setAllowInvalid(false)`
    # refuses the difference IN SILENCE — 30 of 38 rows would have vanished at
    # paste time with no error, on go-live morning.
    #
    # 🔑 Caught only because the importer was repointed at the newer file and
    # the gate ran. The 18 Aug file was read from an Excel export and had ints,
    # so this defect did not exist until the source changed underneath it.
    # ⛔ Normalise here, not at the call site: every caller of split_visa needs it.
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = re.sub(r'\(.*?\)', '', str(raw or '')).strip()
    # Belt and braces for a value that arrived as the STRING '500.0'.
    s = re.sub(r'^(\d+)\.0+$', r'\1', s)
    variant = ''
    m = re.search(r'\b(dependent|subsequent entrant|sponsor|employer)\b', s, re.I)
    if m:
        variant = m.group(1).title()
        s = re.sub(r'\b' + m.group(1) + r'\b', '', s, flags=re.I).strip()
    return re.sub(r'\s+', ' ', s).strip(), variant


def as_date(v):
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date):     return v.isoformat()
    s = str(v or '').strip()
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', s)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        try: return datetime.date(y, mo, d).isoformat()
        except ValueError: return ''
    return ''


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()

    ret, orig = sheets()
    out, held, flags = [], [], collections.Counter()
    notes_extra = collections.defaultdict(list)
    unmatched, seen_email = [], {}

    for row in ret:
        name, visa_raw, team_c, cons, email, party2, skills, last, phone, office, surname = row

        # 1. HELD BACK — the sheet says they are gone. Never create a folder for these.
        if DEAD.search(cons) or DEAD.search(name):
            held.append((name, 'marked "no longer client" in the consultant column'))
            flags['held: no longer a client'] += 1
            continue

        full = name if len(name.split()) > 1 else (name + ' ' + surname).strip()
        if len(full.split()) < 2:
            held.append((name, 'still a single name — no surname supplied'))
            flags['held: no surname'] += 1
            continue

        o = orig.get(norm(name), {})
        if not o:
            unmatched.append(name)
            flags['no matching row in LODGEMENT tab'] += 1

        note = ['Imported from the team\'s returned list on %s (YM-DQ-e573)' % TODAY]

        # 2. Team
        t = TEAM.get(team_c.upper(), '')
        if not t and team_c: note.append('team code %r not F/I' % team_c); flags['team unrecognised'] += 1
        if not team_c:       flags['team BLANK'] += 1

        # 3. Consultant. Two names in one cell -> take the first, record both.
        c_raw = cons.upper().strip()
        if '/' in c_raw:
            first = c_raw.split('/')[0].strip()
            consultant = STAFF.get(first, 'Unassigned')
            note.append('sheet said "%s" — shared file, assigned to the first named' % cons)
            flags['two consultants in one cell'] += 1
        else:
            consultant = STAFF.get(c_raw, '' if not c_raw else 'Unassigned')
            if c_raw and c_raw not in STAFF:
                note.append('consultant column said "%s"' % cons); flags['consultant off-roster'] += 1
        if not c_raw: consultant = 'Unassigned'; flags['consultant BLANK'] += 1

        # 4. Email — validate, never repair
        em = email.strip()
        if em:
            dom = em.split('@')[-1].lower()
            fix = CONFIRMED_FIX['domain'].get(dom)
            if fix:
                em = em[:em.rindex('@') + 1] + fix[0]
                note.append('domain %s -> %s (%s)' % (dom, fix[0], fix[1]))
                flags['email domain corrected — client-confirmed'] += 1
            elif re.match(r'^gm[ai]{0,2}l\.com$', dom) and dom != 'gmail.com':
                # Still refused: looks wrong, but nobody has confirmed THIS one.
                note.append('EMAIL DOMAIN LOOKS WRONG (%s) — confirm before sending' % dom)
                flags['🔴 email domain typo — UNCONFIRMED, not repaired'] += 1
            k = em.lower()
            if k in seen_email:
                note.append('this email is also on another row — confirm which client owns it')
                flags['🔴 duplicate email'] += 1
            seen_email[k] = True
        else:
            flags['email BLANK — no checklist or chase email possible'] += 1

        # 5. Visa
        vt, variant = split_visa(visa_raw)
        if vt and vt not in MAPPED:
            flags['visa has no checklist -> NEEDS REVIEW'] += 1
        if not vt: flags['visa BLANK'] += 1

        # 6. Skills authority — 485 only, and only the five the MAP resolves
        sk = skills.strip()
        # 🔑 CASE-INSENSITIVE FIRST. The client returned "acecqa" in lower case on
        # 25 Aug (D-403). Discarding a correct answer over capitalisation is the
        # same class of loss as accepting a wrong one — and this is a pure
        # normalisation, not a guess about meaning: the letters are identical.
        # ⛔ Only an EXACT case-insensitive hit is resolved. Free text like
        # "Bachelor Degree -no skills assessment" is still blanked and noted,
        # because turning a sentence into a dropdown value IS a judgement.
        if sk and sk not in SKILLS:
            exact = {v.lower(): v for v in SKILLS}.get(sk.lower())
            if exact:
                note.append('skills authority %r normalised to %r (case only)' % (sk, exact))
                flags['skills authority case-normalised'] += 1
                sk = exact
        if sk.lower() in ('n/a', 'na', ''): sk = ''
        elif sk not in SKILLS:
            if '<--' in sk or 'needed' in sk.lower():
                sk = ''
                if vt == '485':
                    note.append('485 with no skills authority — M4 will stamp NEEDS REVIEW')
                    flags['🔴 485 missing skills authority'] += 1
            else:
                note.append('skills authority %r is not one of the five in CHECKLIST MAP' % sk)
                flags['skills authority not a MAP value'] += 1
                sk = ''

        rec = dict.fromkeys(MASTER_HEADERS, '')
        rec['Full Name']           = full
        rec['Email Address']       = em
        if vt in VISA_ALIAS:
            note.append('their sheet said %r; normalised to %r' % (vt, VISA_ALIAS[vt]))
            flags['visa wording normalised to the dropdown'] += 1
            vt = VISA_ALIAS[vt]
        rec['Visa Type']           = vt
        rec['Visa Variant']        = variant
        rec['Office']              = office.upper() if office else 'BRISBANE'
        rec['Team']                = t
        rec['Assigned Consultant'] = consultant
        rec['Skills Authority']    = sk
        rec['Party 2 Name']        = party2
        # 🔴 A CONTACT NUMBER WITH NO DIGITS IS NOT A CONTACT NUMBER. (D-416)
        # Four rows of the 25 Aug return carry the word "OFFSHORE" in column I
        # "7. Contact number" — the team recording onshore/offshore in the phone
        # column. Imported as-is it becomes a phone number, and M7's caller lookup
        # would try to match an incoming call against the string OFFSHORE.
        # ⛔ It also inflated the coverage report: 10 "contact numbers" were really 6.
        if phone and not re.search(r'\d', str(phone)):
            note.append('contact number was %r — not a number, moved to Notes' % phone)
            flags['contact number held text, not a number'] += 1
            phone = ''
        rec['Contact Number']      = phone
        rec['Last Contact']        = as_date(last)
        rec['Visa Expiry']         = as_date(o.get('VISA EXPIRATION'))
        st = str(o.get('STATUS') or '').strip().upper()
        stage, outcome = STAGE.get(st, ('', ''))
        rec['Processing Stage'], rec['Visa Outcome'] = stage, outcome
        if st and st not in STAGE:
            note.append('their STATUS was %r' % st); flags['status word unmapped'] += 1
        if variant: note.append('dependent/secondary applicant — checklist differs')
        # D-352 — Yale already sent every one of these clients their checklist by
        # hand. Stamping the done-marker now means M4 skips them entirely: no file
        # copied, no draft raised. Without it the first run would file 28 checklists
        # and raise 19 drafts nobody asked for.
        rec['Checklist Filed'] = PRESTAMP_CHECKLIST
        rec['Notes'] = ' | '.join(note)
        out.append(rec)

    # ---------------------------------------------------------------- report
    print("=== SOURCE ===")
    print("  rows in the returned list ....... %d" % len(ret))
    print("  HELD BACK ....................... %d" % len(held))
    for n, why in held: print("      · %-28s %s" % (n[:26], why))
    print("  READY TO IMPORT ................. %d" % len(out))
    if unmatched:
        print("  ⚠️  no LODGEMENT row (no status/expiry): %d" % len(unmatched))

    print("\n=== FIELD COVERAGE (of %d importable rows) ===" % len(out))
    for f in ('Team','Assigned Consultant','Email Address','Visa Type','Processing Stage',
              'Visa Expiry','Skills Authority','Contact Number','Last Contact','Party 2 Name'):
        n = sum(1 for r in out if r[f])
        bar = '█' * int(20 * n / max(len(out), 1))
        print("  %-22s %2d/%d %-21s %s" % (f, n, len(out), bar,
              '' if n == len(out) else '← %d missing' % (len(out) - n)))

    print("\n=== WHAT M3 AND M4 WILL DO WITH THIS ===")
    fileable = [r for r in out if r['Visa Type'] in MAPPED and
                (r['Visa Type'] != '485' or r['Skills Authority'])]
    print("  M3 creates a folder for ......... %d  (all of them)" % len(out))
    print("  M4 files a checklist for ........ %d" % len(fileable))
    print("  M4 stamps NEEDS REVIEW for ...... %d" % (len(out) - len(fileable)))
    print("  M4b can draft a checklist email . %d  (needs an email address)"
          % sum(1 for r in fileable if r['Email Address']))
    print("  est. first-run operations ....... ~%d of the 519 left this month"
          % (1 + len(out) * 5 + len(out) * 4))

    print("\n=== FLAGS ===")
    for k, v in flags.most_common(): print("  %-52s %d" % (k, v))

    # ── LOCKED-COLUMN GATE (D-353) ───────────────────────────────────────
    _warn_if_stale(ORIG, "the LODGEMENT tab export", _ORIG_MAX_AGE_DAYS)
    locked = locked_columns()
    bad = collections.defaultdict(list)
    for i, r in enumerate(out, start=1):
        for header, allowed in locked.items():
            v = str(r.get(header, '') or '').strip()
            if v and v not in allowed:
                bad[header].append((i, v))

    print("\n=== LOCKED-COLUMN CHECK — every value against the sheet's own dropdown ===")
    print("  schema read live from %s" % os.path.basename(GS_SETUP))
    for header in sorted(locked):
        n = len(bad.get(header, []))
        print("  %-22s %s" % (header, 'ok' if not n else '🔴 %d value(s) the cell will REJECT' % n))
    if bad:
        print("\n  🔴 THESE ROWS WOULD BE REFUSED AT PASTE TIME, WITH NO ERROR MESSAGE:")
        for header, items in bad.items():
            for i, v in items[:12]:
                print("     row %-3d  %-22s %r" % (i, header, v))
            print("     %-3s  allowed: %s" % ('', ' · '.join(sorted(locked[header]))))

    if a.write and bad:
        # ⛔ Refuse rather than write a CSV that breaks on paste. The whole point of
        # this gate is that the failure it prevents is SILENT — the cell rejects the
        # value and nobody is told which row or why.
        print("\n⛔ REFUSING TO WRITE. Fix the dropdown in setup_master_sheet.gs (then re-run")
        print("   patchMasterDropdowns), or add a VISA_ALIAS entry if it is a wording difference.")
        sys.exit(1)

    if a.write:
        with open(OUT, 'w', newline='', encoding='utf-8') as fh:
            w = csv.DictWriter(fh, fieldnames=MASTER_HEADERS); w.writeheader(); w.writerows(out)
        print("\nWROTE %s  (%d rows)" % (os.path.normpath(OUT), len(out)))
        print("⛔ 40 real people. Outside the repo. Never commit it.")
    else:
        print("\n(report only — pass --write to produce the CSV)")


if __name__ == "__main__":
    main()
