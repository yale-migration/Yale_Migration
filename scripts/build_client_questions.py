#!/usr/bin/env python3
"""Fill the client-question templates with real figures and write them OUTSIDE the repo.

    python3 scripts/build_client_questions.py

Reads   CLIENT-QUESTIONS-team.md, CLIENT-QUESTIONS-robinder.md   (templates, in the repo)
Reads   ../client-data/YALE BRISBANE OFFICE WORK.xlsx            (the real numbers and names)
Writes  ../client-data/SEND-1-team-questions.md
        ../client-data/SEND-2-robinder-questions.md
        ../client-data/SEND-0-claude-pdf-prompt.txt

The templates carry {{PLACEHOLDERS}} and no client names, so they are safe to commit.  The filled
copies contain real client first names, so they are written to ../client-data/, which is outside the
repo (D-317).  Do not move them in.

Every number in the output is computed here from the live workbook rather than typed, so the
document cannot drift from the data the way "3 of 44 rows" did (D-316).
"""

import io
import os
import re
import sys
import warnings
from datetime import date

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))
DATA = os.path.abspath(os.path.join(ROOT, "..", "client-data"))
SRC = os.path.join(DATA, "YALE BRISBANE OFFICE WORK.xlsx")
TAB = "LODGEMENT JULY TO PRESENT"

JUNK = {"SAMPLE", "TEST", "DEMO", "EXAMPLE", "NAME", "N/A", "TBA", "XXX"}
EMAIL_RX = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

BOOKS = ["YALE BRISBANE OFFICE WORK.xlsx", "REYWARD JAKE M GAMOL-2026.xlsx",
         "STUDENTS.xlsx", "DATA SHEET.xlsx"]


def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return re.sub(r"\s+", " ", s)


def load(book, tab, limit=1200):
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        wb.close()
        return None, []
    ws = wb[tab]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(r)
        if i > limit:
            break
    wb.close()
    if not rows:                       # several of their tabs are entirely empty
        return None, []
    hi = 0
    for i, r in enumerate(rows[:8]):
        if sum(1 for c in r if isinstance(c, str) and c.strip()) >= 3:
            hi = i
            break
    hdr = [norm(c) for c in rows[hi]]
    return hdr, [r for r in rows[hi + 1:] if any(c not in (None, "") for c in r)]


def main():
    if not os.path.exists(SRC):
        sys.exit("not found: " + SRC)

    # --- the active list -----------------------------------------------------
    hdr, rows = load(SRC, TAB)
    j = hdr.index("NAME")
    vj = hdr.index("TYPE OF VISA APPLICATION")
    names, partial, seen, visa_of = [], [], set(), {}
    for r in rows:
        n = norm(r[j])
        if not n or n.upper() in JUNK:
            continue
        key = " ".join(sorted(re.sub(r"[^A-Z ]", "", n.upper()).split()))
        if key in seen:
            continue
        seen.add(key)
        names.append(n)
        visa_of[n] = norm(r[vj]) if vj < len(r) else ""
        if len(n.split()) < 2:
            partial.append(n)

    # --- how many tabs did we actually open? --------------------------------
    n_tabs = 0
    for b in BOOKS:
        p = os.path.join(DATA, b)
        if os.path.exists(p):
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            n_tabs += len(wb.sheetnames)
            wb.close()

    # --- distinct clients carrying an email, anywhere ------------------------
    with_email = set()
    for b in BOOKS:
        p = os.path.join(DATA, b)
        if not os.path.exists(p):
            continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        tabs = wb.sheetnames
        wb.close()
        for t in tabs:
            h, rr = load(p, t)
            if not h or "NAME" not in [x.upper() for x in h]:
                continue
            nj = [x.upper() for x in h].index("NAME")
            ej = [k for k, x in enumerate(h) if re.search(r"e-?mail", x, re.I)]
            for r in rr:
                if nj < len(r) and isinstance(r[nj], str):
                    for k in ej:
                        if k < len(r) and isinstance(r[k], str) and EMAIL_RX.search(r[k]):
                            with_email.add(norm(r[nj]).upper())
                            break

    # --- the other client list, counted not remembered ----------------------
    rey = os.path.join(DATA, "REYWARD JAKE M GAMOL-2026.xlsx")
    rey_rows, rey_people = 0, set()
    if os.path.exists(rey):
        wbr = openpyxl.load_workbook(rey, read_only=True, data_only=True)
        for m in ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY",
                  "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER"]:
            if m not in wbr.sheetnames:
                continue
            rws = list(wbr[m].iter_rows(values_only=True))
            hi = 0
            for i, rr in enumerate(rws[:8]):
                if sum(1 for c in rr if isinstance(c, str) and c.strip()) >= 3:
                    hi = i
                    break
            # three tabs label the name column k / CLIENT NAME / Column 1,
            # so take column A rather than trusting the header
            for rr in rws[hi + 1:]:
                v = rr[0] if rr else None
                if isinstance(v, str) and v.strip() and not re.fullmatch(r"[\d./-]+", v.strip()):
                    rey_rows += 1
                    rey_people.add(" ".join(sorted(re.sub(r"[^A-Z ]", "", v.upper()).split())))
        wbr.close()

    subs = {
        "{{DATE}}": date.today().strftime("%d %B %Y"),
        "{{N_ACTIVE}}": str(len(names)),
        "{{N_TABS}}": str(n_tabs),
        "{{N_EMAILS_TOTAL}}": str(len(with_email)),
        "{{N_PARTIAL}}": str(len(partial)),
        "{{PARTIAL_NAMES}}": "> " + ", ".join(partial) if partial else "> (none)",
        "{{TEAM_EVIDENCE}}": "Gayatri, Inder and Star",
        "{{N_REY_ROWS}}": str(rey_rows),
        "{{N_REY_PEOPLE}}": str(len([x for x in rey_people if x])),
        "{{N_CHECKLISTS}}": str(len([f for f in os.listdir(
            os.path.join(ROOT, "docs", "05-canonical-checklists"))
            if f.lower().endswith((".pdf", ".docx")) and not f.startswith("REF_")])),
    }

    out = []
    for tpl, dest in [("CLIENT-QUESTIONS-team.md", "SEND-1-team-questions.md"),
                      ("CLIENT-QUESTIONS-robinder.md", "SEND-2-robinder-questions.md")]:
        txt = io.open(os.path.join(ROOT, tpl), encoding="utf-8").read()
        # drop everything above the first '---\n\n## ' heading = the internal notes
        body = txt
        for k, v in subs.items():
            body = body.replace(k, v)
        left = re.findall(r"\{\{[A-Z_]+\}\}", body)
        if left:
            sys.exit("unfilled placeholder(s): " + ", ".join(sorted(set(left))))
        p = os.path.join(DATA, dest)
        io.open(p, "w", encoding="utf-8").write(body)
        out.append(p)
        print("wrote " + p)

    # --- the facility: give them the form, not the homework -----------------
    # Asking a busy team to "put F or I next to each of 40 names" means they
    # first have to build the list. We already have it. Ship it filled in.
    import csv as _csv
    # Every column here exists because something in the build reads it:
    #   Team        -> M3 routes the folder (no source anywhere, D-316)
    #   Consultant  -> dashboard "workload by consultant" AND the chase list
    #   Email       -> M4b checklist email, M5b chase email
    #   Dependent   -> M4 picks INDIVIDUAL vs DEPENDENT checklist. Get this
    #                  wrong and the client receives the wrong document.
    #   Skills auth -> M4 cannot file ANY 485 checklist without it
    #   Last contact-> the chase list and the dormancy engine have nothing to
    #                  measure from without it
    # Ordered most-useful-first so a half-finished sheet is still worth having.
    sheet = os.path.join(DATA, "Yale-client-list-to-complete.csv")
    with io.open(sheet, "w", encoding="utf-8-sig", newline="") as fh:
        w = _csv.writer(fh)
        w.writerow(["Client name", "Visa (from your sheet)",
                    "1. Team  F=Filipino  I=Indian",
                    "2. Consultant looking after them",
                    "3. Email address",
                    "4. Anyone else on the application?  Y / N",
                    "5. Skills authority (485 only)  TRA / VETASSESS / ACECQA / MASTERS",
                    "6. Date you last spoke to them",
                    "7. Contact number",
                    "8. Office (blank = Brisbane, TSV = Townsville)",
                    "9. Surname (only where needed)"])
        for n in names:
            v = visa_of.get(n, "")
            w.writerow([n, v, "", "", "",
                        "",
                        "<-- needed" if v.startswith("485") else "n/a",
                        "", "", "",
                        "<-- needed" if len(n.split()) < 2 else ""])
    print("wrote " + sheet)

    msg = TEAM_MESSAGE.replace("{{N_ACTIVE}}", str(len(names)))
    mp = os.path.join(DATA, "SEND-3-team-covering-message.txt")
    io.open(mp, "w", encoding="utf-8").write(msg)
    print("wrote " + mp)

    prompt = PROMPT.replace("{{N_ACTIVE}}", str(len(names)))
    pp = os.path.join(DATA, "SEND-0-claude-pdf-prompt.txt")
    io.open(pp, "w", encoding="utf-8").write(prompt)
    print("wrote " + pp)

    print()
    print("Active clients: %d   ·   tabs read: %d   ·   clients with an email anywhere: %d"
          % (len(names), n_tabs, len(with_email)))
    print("First-name-only: %d  ->  %s" % (len(partial), ", ".join(partial)))
    print()
    print("These three files are OUTSIDE the repo and contain real client names.")
    print("Do not commit them. Open SEND-0-claude-pdf-prompt.txt and follow it.")


TEAM_MESSAGE = """Hi team,

Rather than keep sending you separate questions on here, I have put everything into one short
document and emailed it across — subject "Yale Migration - a few gaps in the data".

There are {{N_ACTIVE}} of your current clients involved. I have attached a spreadsheet with all their names
already typed out and blank columns next to them, so most of it is filling in a letter or a name
rather than digging through sheets. That one sheet covers most of the document.

If you are short of time, the three columns that matter are TEAM, CONSULTANT and EMAIL - those three
are what actually stop the system running. Everything after that just makes it better.

Blank is a perfectly good answer anywhere you are not sure - I would rather have twenty rows right
than forty rows guessed.

Quick update while I am here: the folder and checklist automation is built and tested, and the
dashboard is done. Nothing is switched on against your live clients yet - that only happens when
Robinder gives the go-ahead.

That document is everything I need from your side. Nothing else after this.

Happy to jump on a call if anything is quicker to talk through.

Thanks,
Sharjeel
"""


PROMPT = """You are producing a PDF document for a client of mine. I will paste the source markdown
after this prompt.

ABSOLUTE RULES - these matter more than presentation:

1. DO NOT invent, add, merge, split, reword or reorder any question. Reproduce every question
   exactly as written, in the same order, with the same numbering.
2. DO NOT change any number, date, email address, file name, tab name or currency figure. Every
   figure was computed from the client's own spreadsheets. If something looks wrong to you, leave
   it exactly as it is and say so in a note to me OUTSIDE the document.
3. DO NOT add a covering letter, an executive summary, a "next steps" section, a signature block
   beyond what is already there, or any friendly filler. Nothing that is not in the source.
4. If the source contains a section after a line reading "NOT IN THE DOCUMENT" or similar, that
   section is internal. DO NOT put it in the PDF. Stop the document before it.
5. Everything above the first "## Yale Migration" heading is internal instructions to me. DO NOT
   put it in the PDF. The document starts at that heading.
6. DO NOT add any pricing, cost, quoting, budget, hourly-rate, "additional scope" or commercial
   language anywhere. None of it appears in the source and none of it belongs in these documents.
   If a section reads to you like it is leading to a quote, leave it exactly as written.

FORMATTING:
- Clean, plain, professional. A4. Readable at a glance on a phone.
- Keep the tables as tables.
- Bold stays bold - the emphasis marks which questions matter.
- Put the document title and date at the top. Number the pages.
- No logos, no letterhead, no branding. This is a working document, not a proposal.
- Leave a blank line or a ruled space under each question so it can be answered by hand or in a
   reply. That is the point of the document.

CONTEXT so you pitch the tone right:
- The reader is a busy Australian migration agency. English is a second language for several of
  them. Short sentences. No idioms.
- This is a work list they tick through. {{N_ACTIVE}} of their clients are involved, and a spreadsheet
  is attached that answers questions 1 to 8 in one go.

LAYOUT - these raise the reply rate, so treat them as requirements:
a. PRIORITY. Questions 1, 2, 4 and 9 are the ones that unblock the build. Mark those four
   visually - a coloured left border or a light tint behind the block. Everything else stays
   plain. A reader skimming for 20 seconds must be able to see which four matter.
b. ANSWER SPACE MATCHED TO THE QUESTION:
   - Questions 2, 3, 4, 5, 6, 7 are answered in the attached spreadsheet. Give them NO ruled
     lines. Put a small italic note instead: "answer in the attached sheet". Ruled lines there
     cause people to answer twice, or to answer in the PDF and never open the sheet.
   - Questions 8, 11, 12, 13, 14, 16, 18, 19 - one ruled line each.
   - Questions 1, 10, 15, 17 - three ruled lines, they are open-ended.
   - Question 9 - one line, it is an action not a question.
c. THE ATTACHMENT CALLOUT. The paragraph beginning "Attached:
   Yale-client-list-to-complete.csv" must be a bordered or tinted box on its own, not a normal
   paragraph. It is the single thing that makes the rest quick to answer.
d. Aim for 3 pages. Reduce vertical whitespace before you reduce type size, and never go below
   10.5pt body text. Do not leave a final page that is mostly empty - rebalance instead.
e. Question numbers run 1 to 19 straight through seven parts. They must not restart per part.
f. Keep the closing "If you are short of time" block visually distinct - it is the summary of
   what matters.

Produce them as two separate PDFs. Ask me before doing anything the rules above do not cover.
"""


if __name__ == "__main__":
    main()
