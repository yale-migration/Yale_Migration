#!/usr/bin/env python3
"""Exhaustive census of EVERY tab in EVERY client workbook.

    python3 scripts/audit_all_tabs.py

Answers one question and refuses to guess at it:
  does ANY tab, anywhere, carry client EMAIL / OFFICE / TEAM / CONSULTANT?

Why this exists: on 15 Aug we asserted "Office and Team exist in no file they
have sent" after opening four tabs out of roughly ninety.  That is a sample, not
a census, and the client had already been asked a question that depended on it.
A conclusion drawn from a sample must never be reported as a fact about the set.

CREDENTIAL RULE (D-306).  Roughly 1,200 plaintext credentials live in these
workbooks, including ImmiAccount logins and staff phone PINs.  This script:
  * never prints a cell value from a credential column
  * never prints a client email address - only counts and domains
  * skips credential columns entirely when scanning content
"""

import os
import re
import sys
import warnings
from collections import Counter

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.abspath(os.path.join(HERE, "..", "..", "client-data"))

BOOKS = [
    "YALE BRISBANE OFFICE WORK.xlsx",
    "REYWARD JAKE M GAMOL-2026.xlsx",
    "STUDENTS.xlsx",
    "DATA SHEET.xlsx",
]

# NEVER read a column whose header matches this.  Checked before anything else.
BANNED = re.compile(
    r"password|passwd|\buser\s*name\b|\busername\b|\bpwd\b|\botp\b|\bpin\b"
    r"|security\s*question|secret|login\s*detail",
    re.I,
)

WANTED = {
    "EMAIL":      re.compile(r"e-?\s?mail|email\s*add|\bmail\s*id\b", re.I),
    "OFFICE":     re.compile(r"\boffice\b|\bbranch\b|\bcity\b|\blocation\b|brisbane|townsville", re.I),
    "TEAM":       re.compile(r"\bteam\b|filipino|indian|\bnationality\b", re.I),
    "CONSULTANT": re.compile(r"handled\s*by|counsell?or|consultant|assigned|checked\s*by|agent|\bstaff\b|prepared\s*by|\bowner\b", re.I),
    "PHONE":      re.compile(r"contact|phone|mobile|\bcell\b|\bnumber\b", re.I),
}

EMAIL_RX = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

SCAN_ROWS = 400          # deep enough to be representative, bounded for speed


def header_row(rows):
    """First row that looks like headers: 3+ non-empty strings."""
    for i, r in enumerate(rows[:8]):
        if sum(1 for c in r if isinstance(c, str) and c.strip()) >= 3:
            return i
    return 0


def main():
    findings = {k: [] for k in WANTED}
    unlabelled_email = []
    banned_hits = []
    tab_count = 0

    for book in BOOKS:
        path = os.path.join(DATA, book)
        if not os.path.exists(path):
            print("!! missing: " + book)
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print("\n" + "=" * 78)
        print(book + "   (" + str(len(wb.sheetnames)) + " tabs)")
        print("=" * 78)

        for tab in wb.sheetnames:
            tab_count += 1
            ws = wb[tab]
            rows = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                rows.append(r)
                if i >= SCAN_ROWS:
                    break
            if not rows:
                print("  %-32s EMPTY" % tab[:32])
                continue

            hi = header_row(rows)
            hdr = [str(c).strip() if c is not None else "" for c in rows[hi]]
            data = [r for r in rows[hi + 1:] if any(c not in (None, "") for c in r)]

            banned_cols = {j for j, h in enumerate(hdr) if BANNED.search(h)}
            if banned_cols:
                banned_hits.append((book, tab,
                                    [hdr[j] for j in sorted(banned_cols)]))

            hits = []
            for j, h in enumerate(hdr):
                if not h or j in banned_cols:
                    continue
                for kind, rx in WANTED.items():
                    if rx.search(h):
                        fill = sum(1 for r in data
                                   if j < len(r) and r[j] not in (None, ""))
                        if fill:
                            findings[kind].append((book, tab, h, fill, len(data)))
                            hits.append("%s:%s(%d/%d)" % (kind, h[:20], fill, len(data)))
                        break

            # content sweep: an email column with a header we did not predict
            labelled = {j for j, h in enumerate(hdr)
                        if WANTED["EMAIL"].search(h)}
            for j in range(len(hdr)):
                if j in banned_cols or j in labelled:
                    continue
                found = 0
                for r in data:
                    if j < len(r) and isinstance(r[j], str) and EMAIL_RX.search(r[j]):
                        found += 1
                if found >= 2:
                    unlabelled_email.append(
                        (book, tab, hdr[j] or "(col %d, no header)" % (j + 1),
                         found, len(data)))
                    hits.append("EMAIL-IN-CELLS:%s(%d)" % ((hdr[j] or "?")[:16], found))

            print("  %-34s rows=%-5s %s" % (tab[:34], len(data),
                                            "  ".join(hits) if hits else "-"))
        wb.close()

    # ------------------------------------------------------------------ report
    print("\n\n" + "#" * 78)
    print("#  VERDICT   -   " + str(tab_count) + " tabs across " +
          str(len(BOOKS)) + " workbooks, every one opened")
    print("#" * 78)

    for kind in ("EMAIL", "OFFICE", "TEAM", "CONSULTANT", "PHONE"):
        rowset = findings[kind]
        print("\n--- " + kind + " ---")
        if not rowset:
            print("    NOT FOUND in any tab of any workbook.")
            continue
        rowset.sort(key=lambda t: -t[3])
        for book, tab, h, fill, tot in rowset[:14]:
            pct = (100.0 * fill / tot) if tot else 0
            print("    %-26s %-24s %-22s %5d/%-5d  %3.0f%%"
                  % (book[:26], tab[:24], h[:22], fill, tot, pct))
        if len(rowset) > 14:
            print("    ... and %d more" % (len(rowset) - 14))

    print("\n--- EMAIL ADDRESSES FOUND IN CELLS UNDER A NON-EMAIL HEADER ---")
    if not unlabelled_email:
        print("    none")
    else:
        for book, tab, h, n, tot in sorted(unlabelled_email, key=lambda t: -t[3]):
            print("    %-26s %-24s %-22s %5d/%-5d"
                  % (book[:26], tab[:24], h[:22], n, tot))

    print("\n--- CREDENTIAL COLUMNS SKIPPED, NEVER READ (D-306) ---")
    if not banned_hits:
        print("    none seen")
    else:
        for book, tab, cols in banned_hits:
            print("    %-26s %-24s %s" % (book[:26], tab[:24],
                                          ", ".join(c[:24] for c in cols)))
        print("    (%d tabs)" % len(banned_hits))


if __name__ == "__main__":
    main()
